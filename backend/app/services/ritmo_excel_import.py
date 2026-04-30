"""
Importa filas del Excel de facturas (exportación tipo Ritmo / AP).
Columnas esperadas (fila 1): #, Proveedor, RNC, N° Factura, Referencia, Fecha, Vencimiento, Total, Pendiente, Moneda, Estado, Días vencido
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class RitmoRow:
    index: int | None
    proveedor: str
    rnc: str
    n_factura: str
    referencia: str | None
    fecha: date | None
    vencimiento: date | None
    total: Decimal
    pendiente: Decimal | None
    moneda: str
    estado: str
    dias_vencido: int | None


def _to_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-":
        try:
            return date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except ValueError:
            pass
    return None


def _to_dec(val: Any) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace("$", "").replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_ritmo_facturas_excel(
    file_path: Path, *, header_row: int = 1
) -> list[RitmoRow]:
    """Lee la hoja activa; `header_row` es el número de fila de títulos (1-based)."""
    wb = load_workbook(str(file_path), data_only=True)
    sh = wb.active
    if sh is None:
        return []
    # Mapeo por posición (plantilla facturas_ritmo_2026-04-22)
    I = {
        "num": 0,
        "proveedor": 1,
        "rnc": 2,
        "n_fact": 3,
        "ref": 4,
        "fecha": 5,
        "venc": 6,
        "total": 7,
        "pend": 8,
        "mon": 9,
        "est": 10,
        "dv": 11,
    }
    out: list[RitmoRow] = []
    for row in sh.iter_rows(
        min_row=header_row + 1, values_only=True, max_col=12
    ):
        if not any(row):
            continue
        cells = list(row)
        if len(cells) < 8:
            continue

        def c(i: int) -> Any:
            return cells[i] if i < len(cells) else None

        n_fact = str(c(I["n_fact"]) or "").strip()
        if not n_fact:
            continue
        pro = str(c(I["proveedor"]) or "").strip()
        if not pro:
            continue
        total = _to_dec(c(I["total"]))
        if total is None:
            continue
        num_raw = c(I["num"])
        n_int: int | None
        if isinstance(num_raw, (int, float)):
            n_int = int(num_raw) if not isinstance(num_raw, bool) else None
        else:
            s = str(num_raw or "").strip()
            n_int = int(s) if s.isdigit() else None
        d_v = c(I["dv"])
        if isinstance(d_v, (int, float)) and not isinstance(d_v, bool):
            di = int(d_v)
        else:
            t = str(d_v or "").strip()
            try:
                di = int(float(t)) if t else None
            except ValueError:
                di = None

        out.append(
            RitmoRow(
                index=n_int,
                proveedor=pro,
                rnc=str(c(I["rnc"]) or "").strip(),
                n_factura=n_fact,
                referencia=None
                if c(I["ref"]) in (None, "")
                else str(c(I["ref"]))[:200],
                fecha=_to_date(c(I["fecha"])),
                vencimiento=_to_date(c(I["venc"])),
                total=total,
                pendiente=_to_dec(c(I["pend"])),
                moneda=str(c(I["mon"]) or "DOP").strip() or "DOP",
                estado=str(c(I["est"]) or "").strip(),
                dias_vencido=di,
            )
        )
    return out
