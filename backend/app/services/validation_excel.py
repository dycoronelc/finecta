"""Coincidencia de facturas contra Excel del pagador."""
from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db import models
from app.db.models.models import Invoice, InvoiceStatus


@dataclass
class RowInfo:
    invoice_number: str
    amount: Decimal | None
    payer: str | None
    due_date: str | None


def _norm_num(s: str) -> str:
    return re.sub(r"\s+", "", s).upper()


def _find_header_row(sheet) -> int:
    for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
        row_l = " ".join(str(c or "") for c in row).lower()
        if "factura" in row_l or "invoice" in row_l or "monto" in row_l or "importe" in row_l:
            return i
    return 1


def _to_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace("$", "")
    s = s.replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_payer_excel(file_path: Path) -> list[RowInfo]:
    wb = load_workbook(str(file_path), data_only=True)
    sheet = wb.active
    if sheet is None:
        return []
    start = _find_header_row(sheet)
    first = sheet.iter_rows(min_row=start, max_row=start, values_only=True)
    first_row = next(first, None)
    headers = [str(c or "").lower() for c in (first_row or ())]
    idx_inv = 0
    for j, h in enumerate(headers):
        if any(k in h for k in ("invoice", "factura", "ncf", "numero", "número")):
            idx_inv = j
            break
    idx_amount = 0
    for j, h in enumerate(headers):
        if any(
            k in h for k in ("monto", "importe", "total", "amount", "value", "valor")
        ):
            idx_amount = j
            break
    idx_payer = None
    for j, h in enumerate(headers):
        if any(
            k in h for k in ("pagad", "cliente", "deudor", "payer", "debtor", "rnc")
        ):
            idx_payer = j
            break
    idx_due = None
    for j, h in enumerate(headers):
        if any(k in h for k in ("venc", "due", "fecha")):
            idx_due = j
            break
    out: list[RowInfo] = []
    for row in sheet.iter_rows(
        min_row=start + 1, values_only=True, max_col=max(len(headers), 8)
    ):
        if not any(row):
            continue
        cells = list(row)
        inv = str(cells[idx_inv] or "").strip() if idx_inv < len(cells) else ""
        if not inv or inv.lower() in ("totales", "total", "resumen", "ncf", "número"):
            continue
        am = _to_decimal(cells[idx_amount]) if idx_amount < len(cells) else None
        pr = str(cells[idx_payer] or "").strip() if idx_payer is not None and idx_payer < len(cells) else None
        du = str(cells[idx_due] or "") if idx_due is not None and idx_due < len(cells) else None
        out.append(RowInfo(invoice_number=inv, amount=am, payer=pr, due_date=du))
    return out


@dataclass
class MatchItem:
    kind: str
    message: str
    invoice_id: int | None = None
    excel_number: str | None = None


def run_matching(
    db: Session, company_id: int, rows: list[RowInfo]
) -> dict[str, Any]:
    q = select(Invoice).where(Invoice.company_id == company_id)
    invoices = list(db.scalars(q).all())
    by_num: dict[str, list[Invoice]] = {}
    for inv in invoices:
        k = _norm_num(inv.invoice_number)
        by_num.setdefault(k, []).append(inv)
    used_ids: set[int] = set()
    matches: list[MatchItem] = []
    for row in rows:
        k = _norm_num(row.invoice_number)
        found = by_num.get(k) or by_num.get(k.replace(" ", ""))
        if not found:
            matches.append(
                MatchItem(
                    kind="missing_invoice",
                    message="Factura no registrada en el sistema",
                    excel_number=row.invoice_number,
                )
            )
            continue
        inv = found[0]
        if inv.id in used_ids:
            matches.append(
                MatchItem(
                    kind="duplicated",
                    message="Factura duplicada en carga o en sistema",
                    invoice_id=inv.id,
                    excel_number=row.invoice_number,
                )
            )
            continue
        used_ids.add(inv.id)
        if row.payer and inv.payer:
            if _norm_num(row.payer) not in _norm_num(inv.payer) and _norm_num(
                inv.payer
            ) not in _norm_num(row.payer):
                if row.payer[:3].upper() != inv.payer[:3].upper():
                    matches.append(
                        MatchItem(
                            kind="invalid_payer",
                            message="Emisor de pago o pagador no coincide con la factura",
                            invoice_id=inv.id,
                            excel_number=row.invoice_number,
                        )
                    )
        if row.amount and inv.amount:
            if (row.amount - inv.amount).copy_abs() > Decimal("0.5"):
                matches.append(
                    MatchItem(
                        kind="amount_mismatch",
                        message="Monto en Excel no coincide con factura",
                        invoice_id=inv.id,
                        excel_number=row.invoice_number,
                    )
                )
    for inv in invoices:
        if inv.status in (InvoiceStatus.closed, InvoiceStatus.paid, InvoiceStatus.rejected):
            continue
        if inv.id in used_ids:
            continue
        if any(r for r in rows if _norm_num(r.invoice_number) == _norm_num(inv.invoice_number)):
            continue
        if rows:
            matches.append(
                MatchItem(
                    kind="missing_in_excel",
                    message="Factura en sistema no aparece en archivo del pagador",
                    invoice_id=inv.id,
                )
            )
    return {
        "row_count": len(rows),
        "invoices_matched": len(used_ids),
        "items": [asdict(m) for m in matches],
    }
